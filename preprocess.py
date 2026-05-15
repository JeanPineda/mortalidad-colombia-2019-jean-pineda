from collections import Counter
from pathlib import Path
import sys, pickle, pandas as pd
from openpyxl import load_workbook
BASE=Path(__file__).resolve().parent; DATA=BASE/'data'; OUT=DATA/'processed'; OUT.mkdir(exist_ok=True)
STATE=OUT/'state.pkl'; MORT=DATA/'Anexo1.NoFetal2019_CE_15-03-23.xlsx'; DIV=DATA/'Divipola_CE_.xlsx'; COD=DATA/'Anexo2.CodigosDeMuerte_CE_15-03-23.xlsx'
MESES={1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}; SEXO={1:'Hombre',2:'Mujer',3:'Indeterminado'}
def as_code(v,w=None):
    if v is None or pd.isna(v): return ''
    t=str(v).strip(); t=t[:-2] if t.endswith('.0') else t
    return t.zfill(w) if w else t.upper()
def edad(v):
    try: v=int(v)
    except: return 'Edad desconocida'
    return 'Mortalidad neonatal' if 0<=v<=4 else 'Mortalidad infantil' if v<=6 else 'Primera infancia' if v<=8 else 'Niñez' if v<=10 else 'Adolescencia' if v==11 else 'Juventud' if v<=13 else 'Adultez temprana' if v<=16 else 'Adultez intermedia' if v<=19 else 'Vejez' if v<=24 else 'Longevidad / Centenarios' if v<=28 else 'Edad desconocida'
def build_lookup():
    mun=pd.read_excel(DIV, sheet_name='Hoja1', dtype=str)
    for c,w in [('COD_DANE',5),('COD_DEPARTAMENTO',2),('COD_MUNICIPIO',3)]: mun[c]=mun[c].map(lambda x: as_code(x,w))
    mun['DEPARTAMENTO']=mun['DEPARTAMENTO'].str.upper().str.strip(); mun['MUNICIPIO']=mun['MUNICIPIO'].str.upper().str.strip()
    return dict(zip(zip(mun.COD_DANE,mun.COD_DEPARTAMENTO,mun.COD_MUNICIPIO), zip(mun.DEPARTAMENTO,mun.MUNICIPIO)))
def blank(): return {'cnt_depto':Counter(),'cnt_mes':Counter(),'cnt_viol':Counter(),'cnt_mun':Counter(),'cnt_causa':Counter(),'cnt_sexo':Counter(),'cnt_edad':Counter(),'total':0}
def load_state(): return pickle.load(open(STATE,'rb')) if STATE.exists() else blank()
def save_state(st): pickle.dump(st, open(STATE,'wb'))
def process(start,end):
    st=load_state(); lookup=build_lookup(); wb=load_workbook(MORT, read_only=True, data_only=True); ws=wb['No_Fetales_2019']
    for row in ws.iter_rows(min_row=start, max_row=end, values_only=True):
        st['total']+=1
        depto,municipio=lookup.get((as_code(row[0],5),as_code(row[1],2),as_code(row[2],3)),('SIN INFORMACIÓN','SIN INFORMACIÓN'))
        cm=as_code(row[14]); ciudad=f'{municipio} ({depto})'
        st['cnt_depto'][depto]+=1; st['cnt_mes'][MESES.get(row[6],'Sin información')]+=1; st['cnt_mun'][ciudad]+=1; st['cnt_causa'][cm]+=1; st['cnt_sexo'][(depto,SEXO.get(row[9],'Sin información'))]+=1; st['cnt_edad'][edad(row[11])]+=1
        if cm.startswith('X95'): st['cnt_viol'][ciudad]+=1
    save_state(st); print(f'chunk {start}-{end} ok total={st["total"]}')
def export():
    st=load_state()
    # coords
    geo=pd.read_excel(DIV, sheet_name='Hoja3', header=1, dtype=str).rename(columns={'Nombre':'DEPARTAMENTO','Longitud':'LONGITUD','Latitud':'LATITUD'})[['DEPARTAMENTO','LONGITUD','LATITUD']].dropna()
    geo['DEPARTAMENTO']=geo['DEPARTAMENTO'].str.upper().str.strip(); geo['LONGITUD']=pd.to_numeric(geo['LONGITUD'].str.replace(',','.',regex=False),errors='coerce'); geo['LATITUD']=pd.to_numeric(geo['LATITUD'].str.replace(',','.',regex=False),errors='coerce')
    coords=geo.dropna().groupby('DEPARTAMENTO',as_index=False).agg({'LONGITUD':'mean','LATITUD':'mean'})
    cod=pd.read_excel(COD, sheet_name='Final', header=8, dtype=str)
    map4=dict(zip(cod['Código de la CIE-10 cuatro caracteres'].map(as_code), cod['Descripcion  de códigos mortalidad a cuatro caracteres'].astype(str).str.strip()))
    map3=dict(zip(cod['Código de la CIE-10 tres caracteres'].map(as_code), cod['Descripción  de códigos mortalidad a tres caracteres'].astype(str).str.strip()))
    def nom(c):
        c=as_code(c); n=map4.get(c) or map3.get(c[:3]) or 'No especificada'
        return 'No especificada' if str(n).lower()=='nan' else n
    pd.DataFrame([{'DEPARTAMENTO':k,'TOTAL_MUERTES':v} for k,v in st['cnt_depto'].items()]).merge(coords,on='DEPARTAMENTO',how='left').to_csv(OUT/'mapa_departamento.csv',index=False)
    pd.DataFrame([{'MES_NOMBRE':k,'TOTAL_MUERTES':v} for k,v in st['cnt_mes'].items()]).to_csv(OUT/'muertes_mes.csv',index=False)
    pd.DataFrame([{'CIUDAD':k,'TOTAL_HOMICIDIOS':v} for k,v in st['cnt_viol'].items()]).sort_values('TOTAL_HOMICIDIOS',ascending=False).head(5).to_csv(OUT/'top_violencia.csv',index=False)
    pd.DataFrame([{'CIUDAD':k,'TOTAL_MUERTES':v} for k,v in st['cnt_mun'].items() if not k.startswith('SIN INFORMACIÓN')]).sort_values('TOTAL_MUERTES').head(10).to_csv(OUT/'menor_mortalidad.csv',index=False)
    pd.DataFrame([{'COD_MUERTE':k,'NOMBRE_MUERTE':nom(k),'TOTAL_CASOS':v} for k,v in st['cnt_causa'].items()]).sort_values('TOTAL_CASOS',ascending=False).head(10).to_csv(OUT/'causas_top10.csv',index=False)
    pd.DataFrame([{'DEPARTAMENTO':k[0],'SEXO_NOMBRE':k[1],'TOTAL_MUERTES':v} for k,v in st['cnt_sexo'].items()]).to_csv(OUT/'sexo_departamento.csv',index=False)
    pd.DataFrame([{'GRUPO_EDAD_DESC':k,'TOTAL_MUERTES':v} for k,v in st['cnt_edad'].items()]).to_csv(OUT/'grupos_edad.csv',index=False)
    pd.DataFrame([{'TOTAL_REGISTROS':st['total'],'TOTAL_DEPARTAMENTOS':len(st['cnt_depto']),'TOTAL_MUNICIPIOS':len(st['cnt_mun']),'CAUSA_PRINCIPAL':nom(max(st['cnt_causa'].items(),key=lambda x:x[1])[0])}]).to_csv(OUT/'kpis.csv',index=False)
    if STATE.exists(): STATE.unlink()
    print('export ok')
if __name__=='__main__':
    if sys.argv[1]=='process': process(int(sys.argv[2]), int(sys.argv[3]))
    elif sys.argv[1]=='export': export()
